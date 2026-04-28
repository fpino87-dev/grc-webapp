from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from core.audit import log_action
from core.jwt import ExportRateThrottle
from core.scoping import PlantScopedQuerysetMixin

from .models import RiskAppetitePolicy, RiskAssessment, RiskDimension, RiskMitigationPlan
from .serializers import RiskAppetitePolicySerializer, RiskAssessmentSerializer, RiskDimensionSerializer, RiskMitigationPlanSerializer
from .services import get_risk_bia_bcp_context
from .services import delete_risk_assessment


class RiskAssessmentViewSet(PlantScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = RiskAssessment.objects.select_related(
        "plant", "asset", "assessed_by", "accepted_by", "owner", "critical_process"
    )
    serializer_class = RiskAssessmentSerializer
    filterset_fields = ["plant", "status", "assessment_type"]
    plant_field = "plant"

    def destroy(self, request, *args, **kwargs):
        assessment = self.get_object()
        delete_risk_assessment(assessment, request.user)
        return Response(status=204)

    def get_queryset(self):
        qs = super().get_queryset()
        risk_level = self.request.query_params.get("risk_level")
        has_pdca = self.request.query_params.get("has_pdca")
        plant = self.request.query_params.get("plant")

        if plant:
            qs = qs.filter(plant_id=plant)

        if risk_level:
            # Filtra per livello calcolato (verde/giallo/rosso)
            if risk_level == "verde":
                qs = qs.filter(score__lte=7)
            elif risk_level == "giallo":
                qs = qs.filter(score__gt=7, score__lte=14)
            elif risk_level == "rosso":
                qs = qs.filter(score__gt=14)

        if has_pdca == "false":
            from apps.pdca.models import PdcaCycle
            ids_with_pdca = PdcaCycle.objects.exclude(
                fase_corrente="chiuso"
            ).values_list("trigger_source_id", flat=True)
            qs = qs.exclude(pk__in=ids_with_pdca)

        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        log_action(
            user=self.request.user,
            action_code="risk.assessment.create",
            level="L2",
            entity=instance,
            payload={"id": str(instance.id), "assessment_type": instance.assessment_type},
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        # Se l'utente aggiorna campi che cambiano la valutazione,
        # segnaliamo che serve rivalutazione (anche se lo stato non torna a "bozza").
        try:
            from .services import mark_needs_revaluation_if_risk_changed

            changed_fields = set(getattr(serializer, "validated_data", {}) or {}).union(
                set(getattr(self.request, "data", {}) or {})
            )
            mark_needs_revaluation_if_risk_changed(instance, changed_fields)
        except Exception:
            # Non bloccare l'update della UI.
            pass
        log_action(
            user=self.request.user,
            action_code="risk.assessment.update",
            level="L2",
            entity=instance,
            payload={"id": str(instance.id), "status": instance.status},
        )

    @action(detail=True, methods=["post"], url_path="complete")
    def complete(self, request, pk=None):
        from django.utils import timezone
        from .services import calc_ale, calc_score_from_dimensions, escalate_red_risk
        assessment = self.get_object()

        # Calcola score dalle dimensioni IT/OT (con fallback a P×I)
        score = calc_score_from_dimensions(assessment)
        assessment.score = score

        # Calcola ALE automaticamente dalla BIA se collegato
        if assessment.critical_process:
            assessment.ale_annuo = calc_ale(assessment)

        # Genera task se rischio critico
        if score > 14:
            escalate_red_risk(assessment, request.user)

        assessment.status = "completato"
        assessment.assessed_by = request.user
        assessment.assessed_at = timezone.now()
        assessment.save(update_fields=[
            "status", "assessed_by", "assessed_at",
            "score", "ale_annuo", "updated_at",
        ])
        log_action(
            user=request.user,
            action_code="risk.assessment.complete",
            level="L2",
            entity=assessment,
            payload={
                "id": str(assessment.id),
                "score": score,
                "ale_annuo": str(assessment.ale_annuo or 0),
            },
        )
        serializer = self.get_serializer(assessment)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="accept")
    def accept(self, request, pk=None):
        assessment = self.get_object()
        assessment.risk_accepted = True
        assessment.accepted_by = request.user
        assessment.save(update_fields=["risk_accepted", "accepted_by", "updated_at"])
        log_action(
            user=request.user,
            action_code="risk.assessment.accept",
            level="L3",
            entity=assessment,
            payload={"id": str(assessment.id), "risk_accepted": True},
        )
        serializer = self.get_serializer(assessment)
        return Response(serializer.data)


    @action(detail=True, methods=["get"], url_path="suggest-residual")
    def suggest_residual(self, request, pk=None):
        from .services import suggest_residual_score
        assessment = self.get_object()
        return Response(suggest_residual_score(assessment))

    @action(detail=True, methods=["post"], url_path="accept-risk")
    def accept_risk_action(self, request, pk=None):
        from .services import accept_risk
        from django.core.exceptions import ValidationError
        assessment = self.get_object()
        note = request.data.get("note", "")
        expiry_str = request.data.get("expiry_date")
        expiry_date = None
        if expiry_str:
            try:
                from dateutil import parser as dateparser
                expiry_date = dateparser.parse(expiry_str).date()
            except Exception:
                pass
        try:
            accept_risk(assessment, request.user, note, expiry_date)
            return Response({"ok": True})
        except ValidationError as e:
            return Response({"error": str(e.message)}, status=400)


    @action(detail=True, methods=["post"], url_path="renew-acceptance")
    def renew_acceptance(self, request, pk=None):
        """
        Rinnova la scadenza dell'accettazione formale.
        Body: { "expiry_date": "YYYY-MM-DD" }  — default: oggi + 1 anno.
        Resetta needs_revaluation=False se era stato alzato per scadenza.
        """
        from django.utils import timezone
        from dateutil import parser as dateparser

        assessment = self.get_object()
        if not assessment.risk_accepted_formally:
            return Response({"error": "Il rischio non è ancora accettato formalmente."}, status=400)

        expiry_str = request.data.get("expiry_date")
        if expiry_str:
            try:
                expiry_date = dateparser.parse(expiry_str).date()
            except Exception:
                return Response({"error": "Data non valida."}, status=400)
        else:
            expiry_date = timezone.now().date() + timezone.timedelta(days=365)

        today = timezone.now().date()
        if expiry_date <= today:
            return Response({"error": "La nuova scadenza deve essere successiva a oggi."}, status=400)

        assessment.risk_acceptance_expiry = expiry_date
        assessment.needs_revaluation = False
        assessment.needs_revaluation_since = None
        assessment.save(update_fields=[
            "risk_acceptance_expiry", "needs_revaluation", "needs_revaluation_since", "updated_at"
        ])
        log_action(
            user=request.user,
            action_code="risk.acceptance.renewed",
            level="L2",
            entity=assessment,
            payload={"id": str(assessment.id), "new_expiry": str(expiry_date)},
        )
        return Response(self.get_serializer(assessment).data)

    @action(detail=True, methods=["post"], url_path="reset-acceptance")
    def reset_acceptance(self, request, pk=None):
        """
        Annulla l'accettazione formale per rivalutazione anticipata.
        Setta risk_accepted_formally=False e needs_revaluation=True.
        """
        from django.utils import timezone

        assessment = self.get_object()
        assessment.risk_accepted_formally = False
        assessment.risk_accepted_by = None
        assessment.risk_accepted_at = None
        assessment.risk_acceptance_note = ""
        assessment.risk_acceptance_expiry = None
        assessment.needs_revaluation = True
        assessment.needs_revaluation_since = timezone.now().date()
        assessment.save(update_fields=[
            "risk_accepted_formally", "risk_accepted_by", "risk_accepted_at",
            "risk_acceptance_note", "risk_acceptance_expiry",
            "needs_revaluation", "needs_revaluation_since", "updated_at",
        ])
        log_action(
            user=request.user,
            action_code="risk.acceptance.reset",
            level="L3",
            entity=assessment,
            payload={"id": str(assessment.id), "reset_by": request.user.email},
        )
        return Response(self.get_serializer(assessment).data)

    @action(detail=False, methods=["get"], url_path="needs-revaluation")
    def needs_revaluation_list(self, request):
        """Risk assessment che richiedono rivalutazione dopo un change."""
        plant_id = request.query_params.get("plant")
        qs = self.get_queryset().filter(needs_revaluation=True)
        if plant_id:
            qs = qs.filter(plant_id=plant_id)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=True, methods=["get"], url_path="context")
    def context(self, request, pk=None):
        """
        Vista di contesto per un singolo risk assessment:
        include BIA del processo collegato e BCP che lo coprono.
        """
        assessment = self.get_object()
        data = get_risk_bia_bcp_context(assessment)
        return Response(data)

    @action(
        detail=False, methods=["get"], url_path="export",
        throttle_classes=[ExportRateThrottle],
    )
    def export(self, request):
        """
        Export Excel del registro rischi filtrato per plant.
        Solo rischi completati di default; ?include_draft=1 per includere bozze.
        """
        import io
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from .models import NIS2_ART21_CHOICES, NIS2_RELEVANCE_CHOICES

        from django.db.models import Prefetch
        plant_id = request.query_params.get("plant")
        include_draft = request.query_params.get("include_draft") == "1"

        # Prefetch mitigation plans con owner in una sola query
        plans_prefetch = Prefetch(
            "mitigation_plans",
            queryset=RiskMitigationPlan.objects.select_related("owner").order_by("due_date"),
        )
        qs = (
            RiskAssessment.objects.select_related(
                "plant", "owner", "risk_accepted_by", "critical_process"
            )
            .prefetch_related(plans_prefetch)
            .filter(deleted_at__isnull=True)
        )
        if plant_id:
            qs = qs.filter(plant_id=plant_id)
        if not include_draft:
            qs = qs.filter(status="completato")

        from .services import calc_ale

        art21_map = dict(NIS2_ART21_CHOICES)
        relevance_map = dict(NIS2_RELEVANCE_CHOICES)
        prob_labels = {1: "Molto bassa", 2: "Bassa", 3: "Media", 4: "Alta", 5: "Molto alta"}
        impact_labels = {1: "Trascurabile", 2: "Minore", 3: "Moderato", 4: "Grave", 5: "Critico"}
        level_labels = {"verde": "Basso", "giallo": "Medio", "rosso": "Alto/Critico"}
        treatment_labels = {
            "mitigare": "Mitigare", "accettare": "Accettare",
            "trasferire": "Trasferire", "evitare": "Evitare",
        }

        headers = [
            # Descrizione
            "Nome / Scenario", "Causa", "Conseguenza",
            # Classificazione
            "Tipo (IT/OT)", "Categoria minaccia", "Owner",
            # Scoring inerente
            "Prob. inerente", "Impatto inerente", "Score inerente",
            # Scoring residuo
            "Probabilità residua", "Impatto residuo", "Score residuo", "Livello rischio",
            # ALE
            "ALE (€)",
            # Trattamento
            "Trattamento", "Scadenza piano",
            # Azioni di mitigazione (da MitigationPlan)
            "Azioni di mitigazione",
            # Accettazione formale Art.20
            "Rischio accettato (Art.20)", "Accettato da", "Data accettazione",
            "Scadenza accettazione", "Nota accettazione",
            # NIS2
            "Sistemi impattati (NIS2)", "Art.21 NIS2", "Rilevanza NIS2",
            # Contesto
            "Processo BIA", "Plant", "Stato",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Register"

        header_fill = PatternFill("solid", fgColor="1E3A5F")
        accept_fill = PatternFill("solid", fgColor="1A5276")
        nis2_fill = PatternFill("solid", fgColor="0F5E3A")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        nis2_cols = {headers.index(h) for h in ("Sistemi impattati (NIS2)", "Art.21 NIS2", "Rilevanza NIS2")}
        accept_cols = {headers.index(h) for h in ("Rischio accettato (Art.20)", "Accettato da", "Data accettazione", "Scadenza accettazione", "Nota accettazione")}

        for col_idx, header in enumerate(headers, 1):
            i = col_idx - 1
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = nis2_fill if i in nis2_cols else (accept_fill if i in accept_cols else header_fill)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        level_colors = {"verde": "C6EFCE", "giallo": "FFEB9C", "rosso": "FFC7CE"}

        for row_idx, risk in enumerate(qs.order_by("created_at"), 2):
            def _name(u):
                if not u:
                    return ""
                return f"{u.first_name} {u.last_name}".strip() or u.email

            # Piani di mitigazione: concatenati con newline
            plans_text = ""
            for p in risk.mitigation_plans.all():
                plan_owner = _name(p.owner)
                status_str = "✓ Completato" if p.completed_at else "In corso"
                line = f"• {p.action}"
                if plan_owner:
                    line += f" [{plan_owner}]"
                line += f" — Scad: {p.due_date} — {status_str}"
                plans_text += line + "\n"
            plans_text = plans_text.strip()

            # ALE: usa quello calcolato dalla BIA se disponibile, altrimenti annuo
            ale_val = calc_ale(risk)
            ale_str = str(ale_val) if ale_val else (str(risk.ale_annuo) if risk.ale_annuo else "")

            row = [
                risk.name,
                risk.cause,
                risk.consequence,
                risk.assessment_type,
                risk.get_threat_category_display() if risk.threat_category else "",
                _name(risk.owner),
                prob_labels.get(risk.inherent_probability, ""),
                impact_labels.get(risk.inherent_impact, ""),
                risk.inherent_score or "",
                prob_labels.get(risk.probability, ""),
                impact_labels.get(risk.impact, ""),
                risk.score or "",
                level_labels.get(risk.risk_level, ""),
                ale_str,
                treatment_labels.get(risk.treatment, risk.treatment),
                str(risk.plan_due_date) if risk.plan_due_date else "",
                plans_text,
                "Sì" if risk.risk_accepted_formally else ("In attesa" if risk.treatment == "accettare" else "No"),
                _name(risk.risk_accepted_by),
                str(risk.risk_accepted_at.date()) if risk.risk_accepted_at else "",
                str(risk.risk_acceptance_expiry) if risk.risk_acceptance_expiry else "",
                risk.risk_acceptance_note,
                risk.impacted_systems,
                art21_map.get(risk.nis2_art21_category, ""),
                relevance_map.get(risk.nis2_relevance, ""),
                risk.critical_process.name if risk.critical_process else "",
                risk.plant.name if risk.plant else "",
                risk.status,
            ]

            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            level_col = headers.index("Livello rischio") + 1
            color = level_colors.get(risk.risk_level)
            if color:
                ws.cell(row=row_idx, column=level_col).fill = PatternFill("solid", fgColor=color)

        # Larghezze colonne (una per header)
        col_widths = [
            35, 40, 40,          # nome, causa, conseguenza
            10, 22, 20,          # tipo, categoria, owner
            16, 16, 12,          # prob/imp/score inerente
            16, 16, 12, 14,      # prob/imp/score/livello residuo
            14,                  # ALE
            12, 14,              # trattamento, scadenza piano
            50,                  # azioni mitigazione
            18, 20, 18, 16, 40,  # accettazione Art.20
            35, 35, 22,          # NIS2
            22, 20, 12,          # processo BIA, plant, stato
        ]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from apps.plants.models import Plant
        plant_obj = Plant.objects.filter(pk=plant_id).first() if plant_id else None
        if plant_obj:
            log_action(
                user=request.user,
                action_code="risk.assessment.export",
                level="L2",
                entity=plant_obj,
                payload={"plant_id": str(plant_id), "count": qs.count()},
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        plant_label = f"_plant_{plant_id}" if plant_id else ""
        filename = f"risk_register{plant_label}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class RiskDimensionViewSet(PlantScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = RiskDimension.objects.select_related("assessment")
    serializer_class = RiskDimensionSerializer
    filterset_fields = ["plant"]
    plant_field = "assessment__plant"

    def get_queryset(self):
        qs = super().get_queryset()
        plant = self.request.query_params.get("plant")
        if plant:
            qs = qs.filter(assessment__plant_id=plant)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        log_action(
            user=self.request.user,
            action_code="risk.dimension.create",
            level="L1",
            entity=instance,
            payload={"id": str(instance.id), "dimension_code": instance.dimension_code},
        )


class RiskMitigationPlanViewSet(PlantScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = RiskMitigationPlan.objects.select_related("assessment", "owner", "control_instance")
    serializer_class = RiskMitigationPlanSerializer
    filterset_fields = ["assessment"]
    plant_field = "assessment__plant"

    def get_queryset(self):
        qs = super().get_queryset()
        plant = self.request.query_params.get("plant")
        assessment = self.request.query_params.get("assessment")
        if plant:
            qs = qs.filter(assessment__plant_id=plant)
        if assessment:
            qs = qs.filter(assessment_id=assessment)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        log_action(
            user=self.request.user,
            action_code="risk.mitigation_plan.create",
            level="L2",
            entity=instance,
            payload={"id": str(instance.id)},
        )

    def perform_update(self, serializer):
        instance = serializer.save()
        log_action(
            user=self.request.user,
            action_code="risk.mitigation_plan.update",
            level="L2",
            entity=instance,
            payload={"id": str(instance.id), "completed_at": str(instance.completed_at)},
        )


class RiskAppetitePolicyViewSet(PlantScopedQuerysetMixin, viewsets.ModelViewSet):
    queryset = RiskAppetitePolicy.objects.select_related("plant", "approved_by")
    serializer_class = RiskAppetitePolicySerializer
    filterset_fields = ["plant", "framework_code"]
    plant_field = "plant"
    allow_null_plant = True  # plant=null = policy org-wide valida per tutti i plant

    def perform_create(self, serializer):
        instance = serializer.save(created_by=self.request.user)
        log_action(
            user=self.request.user,
            action_code="risk.appetite_policy.create",
            level="L1",
            entity=instance,
            payload={
                "max_acceptable_score": instance.max_acceptable_score,
                "framework_code": instance.framework_code,
            },
        )

    @action(detail=False, methods=["get"], url_path="active")
    def active(self, request):
        """Recupera la policy attiva per plant e framework."""
        from .services import get_active_appetite
        plant_id = request.query_params.get("plant")
        framework_code = request.query_params.get("framework", "")
        from apps.plants.models import Plant
        plant = Plant.objects.filter(pk=plant_id).first() if plant_id else None
        policy = get_active_appetite(plant=plant, framework_code=framework_code)
        if not policy:
            return Response({"detail": "Nessuna policy attiva"}, status=404)
        return Response(RiskAppetitePolicySerializer(policy).data)
